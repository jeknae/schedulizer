from itertools import product
from typing import BinaryIO

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet


Column = int


def load_worksheet(file: str | BinaryIO) -> Worksheet:
    """Load worksheet
    :param file: the path to open file or file-like object

    :return: `openpyxl.worksheet.worksheet.Worksheet`
    """
    wb = load_workbook(file)
    sheets = wb.sheetnames
    ws = wb[sheets[0]]
    return ws


class BaseExcellScrapper:
    def __init__(self, file: str | BinaryIO):
        self.ws = load_worksheet(file)

    def __find_value(self,
                     value: str,
                     f_range: range,
                     s_range: range):

        for row, col in product(f_range, s_range):
            if value == str(self.ws.cell(row, col).value).lower().strip():
                return self.ws.cell(row, col)
        return self.ws.cell(1, 1)

    def __find_type(self,
                    _type: type,
                    f_range: range,
                    s_range: range):
        for row, col in product(f_range, s_range):
            if issubclass(type(self.ws.cell(row, col).value), _type):
                return self.ws.cell(row, col)
        return self.ws.cell(1, 1)

    def _get_cell_by_value(self,
                           value: str,
                           row_range: range,
                           column_range: range,
                           revert: bool = False) -> Cell:
        """
        Return first cell that contains needed value otherwise return cell in 1 row and 1 column

        :param value: value that must be in cell
        :param row_range: row range within which must be produced search
        :param column_range: column range witch which must be produced search
        :param revert: order for search (rows by column: revert=True or columns by row: revert=False)
        :return: :class: `openpyxl.cell.cell.Cell`
        """
        value = value.lower().strip()

        if revert:
            return self.__find_value(value, column_range, row_range)
        else:
            return self.__find_value(value, row_range, column_range)

    def _get_cell_by_type(self,
                          _type: type,
                          row_range: range,
                          column_range: range,
                          revert: bool = False) -> Cell:
        """
        Return first cell that contains value with needed type otherwise return cell in 1 row and 1 column

        :param _type: type that must be in cell
        :param row_range: row range within which must be produced search
        :param column_range: column range witch which must be produced search
        :return: :class: `openpyxl.cell.cell.Cell`
        """
        if revert:
            return self.__find_type(_type, column_range, row_range)
        else:
            return self.__find_type(_type, row_range, column_range)
