from datetime import datetime
from typing import BinaryIO

from openpyxl.cell.cell import Cell

# user imports
from schedulizer.excell_scrapper.core.base_scrapper import BaseExcellScrapper


Column = int


class IIGUExcellScrapper(BaseExcellScrapper):
    """Class for IIGU excell scrapper

    Schedule places in:
        row: range(1, 100)
        column: range(1, 20)

        Group name:
            row: range(1, 10)
            column: range(1, 20)
        Date:
            row: range(1, 10)
            column: range(1, 5)
        Week day:
            row: in one row with date
            column: under each group name
        Time:
            row range: between two dates, days, or in each line with lesson
            column: date column
        Lesson:
            row: between two days, dates, or in each line with time
            column: under group name

    """
    def __init__(self, file: str | BinaryIO, group_name: str | None = None):
        super().__init__(file)
        self.group_name = group_name

    def get_group_names(self) -> list[str]:
        """Return all groups name"""
        cell = self._get_cell_by_value('название групп', range(1, 20), range(1, 4))
        row = cell.row
        column_start = cell.column+1
        groups = []
        for column in range(column_start, 20):
            name = self.ws.cell(row, column).value
            if name is not None:
                groups.append(name)
        return groups

    def _get_group_cell(self) -> Cell:
        """Return group column
        :return: `openpyxl.cell.cell.Cell`
        """
        return self._get_cell_by_value(self.group_name, range(1, 10), range(1, 20))

    def _get_dates_column(self) -> Column:
        """Return column of dates
        :return: int
        """
        return self._get_cell_by_type(datetime, range(1, 10), range(1, 5), revert=True).column

    def _get_date_cell(self, week_day: str) -> Cell:
        """Return needed week day cell

            Under every group name is stored a week day cell

            Each date is stored in one row with week day

        :return: :class: `openpyxl.cell.cell.Cell`
        """
        date_column = self._get_dates_column()
        group_column = self._get_group_cell().column
        row = self._get_cell_by_value(value=week_day,
                                      row_range=range(1, 100),
                                      column_range=range(group_column, group_column+1)).row
        return self.ws.cell(row, date_column)

    def _get_day_row_range(self, week_day: str) -> range:
        """Return range of rows for one day

            range-start value is a day row

            range-end value is a next day row

            sunday range-end is sum of sunday row number and 10
        :return: range
        """
        week = {1: "Понедельник", 2: "Вторник",
                3: "Среда", 4: "Четверг",
                5: "Пятница", 6: "Суббота"}
        invert = {i: j for j, i in week.items()}

        day = int(invert.get(week_day))

        if week_day == "Суббота":
            return range(self._get_date_cell(week_day).row + 1, self._get_date_cell(week[6]).row + 10)
        return range(self._get_date_cell(week[day]).row + 1, self._get_date_cell(week[day + 1]).row)

    def _get_lesson_cells(self, week_day: str) -> list[Cell]:
        """Return lesson cells

            If cells is vertical merged then its value is stored in upper cell

            If cells is horizontal merged then its value is stored in left cell

            therefore need to do shift (row-1) (column-1).
        :return: list[openpyxl.cell.cell.Cell]
        """
        column = self._get_group_cell().column
        rows = self._get_day_row_range(week_day)
        lessons = []
        for row in rows:
            cell = self.ws.cell(row, column)
            if type(cell).__name__ == "Cell" and cell.value is not None:
                lessons.append(cell)
            elif type(cell).__name__ == "MergedCell":
                if type(self.ws.cell(row - 1, column)).__name__ == "MergedCell":
                    if self.ws.cell(row - 1, column - 1).value is not None:
                        lessons.append(self.ws.cell(row - 1, column - 1))

        return lessons

    def _get_time_cells(self, week_day: str) -> list[Cell]:
        """Return list of time cells

            Time row is stored in one line with lesson row

            Time column is stored in one column with date

        :return: tuple[openpyxl.cell.cell.Cell]
        """
        column = self._get_dates_column()
        cells = self._get_lesson_cells(week_day)
        if len(cells) > 0:
            return [self.ws.cell(cell.row, column) for cell in cells]
        return []

    def get_lessons(self, week_day: str) -> list[str]:
        """Return list of lessons

        :return: tuple[str, ...]
        """
        cells = self._get_lesson_cells(week_day)
        if cells is not None:
            return [f'{cell.value}' for cell in cells]
        return []

    def get_times(self, week_day: str) -> list[str]:
        """Return list of times
        :return: list[str]
        """
        cells = self._get_time_cells(week_day)
        if cells is not None:
            return [f"{i.value.replace(' ', '').replace('.', ':')}" for i in cells]
        return []

    def get_date(self, week_day: str) -> str:
        """Return date
        :return: str
        """
        return self._get_date_cell(week_day).value
